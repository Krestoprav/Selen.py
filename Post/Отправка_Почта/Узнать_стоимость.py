# -*- coding: utf-8 -*-
# примеры работы с api v1.0 для pyton 3
# необходим установленный модуль "requests"
# документация модуля: https://requests.readthedocs.io
# инструкции по установке модуля: https://requests.readthedocs.io/en/master/user/install

import requests
import json
import telebot
from selenium import webdriver
from selenium.webdriver.chrome.webdriver import WebDriver
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoSuchElementException, TimeoutException
import time
from time import sleep
import pyperclip
import numpy
import datetime
from datetime import timedelta, datetime
import sys
import os
import re
import shutil
import httplib2
import apiclient.discovery
import gspread  # импортируем gspread
from oauth2client.service_account import ServiceAccountCredentials  # ипортируем ServiceAccountCredentials
from zeep import CachingClient, Settings, helpers
import openpyxl
from openpyxl import Workbook, load_workbook, formatting



def send_type():
    """Функция проставляет полученные из API типы отправлений в гуглтаблицы"""

    driver_link=rf'C:\Users\User\Desktop\Python\Drivers\chromedriver.exe'  # Ссылка на драйвер

    num_col_send_type=21  # Номер столбца в гуглтаблице, в который будем вносить тип отправления

    # Блок запуска гугл таблицы
    link = ['https://spreadsheets.google.com/feeds',
            'https://www.googleapis.com/auth/drive']  # задаем ссылку на Гугл таблици
    my_creds = ServiceAccountCredentials.from_json_keyfile_name(
        rf'C:\Users\User\Desktop\Python\Drivers\krestoprav-8e5f49dc5ebe.json',
        link)  # формируем данные для входа из нашего json файла
    client = gspread.authorize(my_creds)  # запускаем клиент для связи с таблицами

    # Определяем проект, с которым будем работать
    # Открываем файл Эксель с проектами
    tab_name=r'C:\Users\User\Desktop\Python\FIO_sender\Projects.xlsx'
    tab = load_workbook(tab_name)
    # Открываем нужный лист в файле
    sheet_xlsx = tab['Лист1']  # Открываем вкладку с проектами

    # Получаем список последних дат заполнения ФИО в проекты
    listE = []
    for col in sheet_xlsx['E']:
        # print(col.value)
        if col.value == None:
            continue
        listE.append(col.value)

    # Получаем список проектов
    listB = []
    for col in sheet_xlsx['B']:
        # print(col.value)
        if col.value == None:
            continue
        listB.append(col.value)

    print(*listE, sep='\n')  # Печатаем список дат
    print('Самая старая дата - ' + str(min(listE)))  # Печатаем самую старую дату из списка
    ind = int(listE.index(min(listE))) + 1  # Получаем номер строки с нужной датой (к индексу из списка прибавляем +1)
    URL = str(sheet_xlsx[f'C{ind}'].value)  # Получаем ссылку на гуглтаблицу проекта
    vkl = str(sheet_xlsx[f'F{ind}'].value)  # Получаем название вкладки из гуглтаблицы
    num_col_ind = str(sheet_xlsx[f'G{ind}'].value)  # Получаем номер колонки с индексом в гугл таб
    print(URL)
    print(rf"Выбираем проект '{sheet_xlsx[f'B{ind}'].value}', последняя дата его заполнения {min(listE)}")

    sheet = client.open_by_url(URL).worksheet(str(vkl))  # открываем нужную нам таблицу и лист


    # Определяем количество значений в колонке с типом отправления отправителя
    n = len(sheet.col_values(num_col_send_type))
    # n=10261
    # Печатаем список отправителей из колонки (для самопроверки)
    # print(sheet.col_values(num_col_send_type))
    # Если в колонке нет ни одного отправителя, то начинаенм со второй строки
    if n==0:
        n=2
    # Иначе - начинаем работать со следующей строки после последней заполненной
    else:
        n+=1
    print(rf'Начинаем работать со троки {n}')

    # Цикл обхода строк вкладки, пока первый столбец не пустой
    while sheet.cell(n, 1).value != None:
        # Цикл обхода ошибок при определении столбца с типом отправления и статусом
        while True:
            try:
                a=sheet.cell(n, num_col_send_type).value
                stat=sheet.cell(n, 2).value
            except Exception:
                print("Ошибка при определении статуса и наличия типа отправления")
                time.sleep(30)
                continue
            break
        # Если типа отправления нет а трек есть
        if a == None and (stat == 'Подтвержден' or stat == 'подтверждён'):
            print(rf"{n} {stat} заполняем")
            # Цикл обхода ошибок при получении индекса
            while True:
                try:
                    index = sheet.cell(n, num_col_ind).value  # Получаем индекс
                    print(index)
                except Exception:
                    print("Ошибка при получении индекса из таблицы")
                    time.sleep(15)
                    continue
                break

            def api(index):
                """Функция для запроса API"""

                global RPOrus
                RPOrus = ["Обычные", "Онлайн", "Онлайн курьер"]
                RPOs = ["POSTAL_PARCEL", "ONLINE_PARCEL", "ONLINE_COURIER"]
                coasts = []

                for RPO in RPOs:
                    # properties
                    protocol = "https://"
                    host = "otpravka-api.pochta.ru"
                    token = "zpRNYASJazFUNLCXD8gO1t5f4UM08OEn"
                    key = "b3MucHJldGVueml5QHlhbmRleC5ydTptb2lzZWV2YTIxMTE="

                    request_headers = {
                        "Content-Type": "application/json",
                        "Accept": "application/json;charset=UTF-8",
                        "Authorization": "AccessToken " + token,
                        "X-User-Authorization": "Basic " + key
                    }

                    path = "/1.0/tariff"

                    destination = {
                        "index-from": "141020",
                        "index-to": str(index),
                        "mail-category": "WITH_DECLARED_VALUE_AND_CASH_ON_DELIVERY",
                        "mail-type": RPO,
                        "mass": 500,
                        "declared-value": 3300,
                        "fragile": "false"
                    }

                    url = protocol + host + path

                    response = requests.post(url, headers=request_headers, data=json.dumps(destination))
                    # print("Status code: ", response.status_code)
                    print("Response body: ", response.json())
                    coast = (float(response.json()['total-rate']) + float(
                        response.json()['total-rate'])) / 100
                    if coast==0.0:
                        coast=100000
                    print(rf"{RPO} {coast}")
                    coasts.append(coast)

                print(coasts)
                a = min(coasts)
                print("Минимальная цена - " + str(a))
                w = coasts.index(a)
                print("Положение в списке - " + str(w))
                print(RPOrus[w])
                global send_t
                send_t = RPOrus[w]

            # Три раза пробуем определить тип отправления
            for q in range(3):
                try:
                    api(index)
                except Exception as E:
                    print(E.args)
                    continue
                break

            # print(rf"Печатаем тип отправления {send_t}")

            # цикл обхода ошибки при вставке типа отправления в таблицу
            while True:
                try:
                    sheet.update_cell(n, num_col_send_type, send_t)
                except Exception:
                    print("Ошибка при вставке индекса в таблицу")
                    time.sleep(15)
                    continue
                break
        else:
            print(rf"{n} пропускаем")
        n+=1
        time.sleep(5)


    print("Список закончился")


    """Заносим последнюю дату в файл Эксель"""
    sheet_xlsx[f'E{ind}'] = datetime.now()
    tab.save(tab_name)
    print("Файл xlsx сохранен")


if __name__ == '__main__':

    while True:
        try:
            send_type()
        except Exception as E:
            print(E.args)
            time.sleep(30)
            continue
