import telebot
from selenium import webdriver
from selenium.webdriver.chrome.webdriver import WebDriver
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoSuchElementException, TimeoutException
import time
from time import sleep
import pyperclip
import numpy
import datetime
from datetime import timedelta, datetime
import sys
import os
import re
import shutil
import httplib2
import apiclient.discovery
import gspread  # импортируем gspread
from oauth2client.service_account import ServiceAccountCredentials  # ипортируем ServiceAccountCredentials
from zeep import CachingClient, Settings, helpers
import openpyxl
from openpyxl import Workbook, load_workbook, formatting

def FIO():

    driver_link=rf'C:\Users\User\Desktop\Python\Drivers\chromedriver.exe'

    # Блок запуска гугл таблицы
    link = ['https://spreadsheets.google.com/feeds',
            'https://www.googleapis.com/auth/drive']  # задаем ссылку на Гугл таблици
    my_creds = ServiceAccountCredentials.from_json_keyfile_name(
        rf'C:\Users\User\Desktop\Python\Drivers\krestoprav-8e5f49dc5ebe.json',
        link)  # формируем данные для входа из нашего json файла
    client = gspread.authorize(my_creds)  # запускаем клиент для связи с таблицами

    # Определяем проект, с которым будем работать
    # Открываем файл Эксель с проектам
    tab_name=r'C:\Users\User\Desktop\Python\FIO_sender\Projects.xlsx'
    tab = load_workbook(tab_name)
    # Открываем нужный лист в файле
    sheet_xlsx = tab['Лист1']  # Открываем вкладку с проектами

    # Получаем список последних дат заполнения ФИО в проекты
    listD = []
    for col in sheet_xlsx['D']:
        # print(col.value)
        if col.value == None:
            continue
        listD.append(col.value)

    # Получаем список проектов
    listB = []
    for col in sheet_xlsx['B']:
        # print(col.value)
        if col.value == None:
            continue
        listB.append(col.value)

    print(*listD, sep='\n')  # Печатаем список дат
    print('Самая старая дата - ' + str(min(listD)))  # Печатаем самую старую дату из списка
    index = int(listD.index(min(listD)))+1  # Получаем номер строки с нужной датой (к индексу из списка прибавляем +1)
    URL = str(sheet_xlsx[f'C{index}'].value)
    print(URL)
    print(rf"Выбираем проект '{sheet_xlsx[f'B{index}'].value}', последняя дата его заполнения {min(listD)}")


    # URL ="https://docs.google.com/spreadsheets/d/1V8_RdqLsLwPAS_pNlOBQHCEa3pRf_lraTbhp7yEpqCs/edit#gid=715714864"

    sheet = client.open_by_url(URL).worksheet('Ответы на форму (1)')  # открываем нужную нам таблицу и лист

    #  Блока запуска браузера

    options = webdriver.ChromeOptions()
    options.add_argument('headless')
    options.add_argument('window-size=1920x935')

    browser = webdriver.Chrome(driver_link, options=options)
    browser.maximize_window()
    wait = WebDriverWait(browser, 30)

    # Первая строка в таблицу, с которой начинается обход (чтение из файла)
    # file = open('1.txt', 'r', encoding='utf-8')
    # n=int(file.readline())
    # file.close()

    # Определяем количство значений в колонке с ФИО отправителя
    n = len(sheet.col_values(18))

    # Печатаем список отправителей из колонки (для самопроверки)
    # print(sheet.col_values(18))
    # Если в колонке нет ни одного отправителя, то начинаенм со второй строки
    if n==0:
        n=2
    # Иначе - начинаем работать со следующей строки после последней заполненной
    else:
        n+=1
    print(rf'Начинаем работать со троки {n}')

    # Цикл обхода строк вкладки, пока первый столбец не пустой
    while sheet.cell(n, 1).value != None:
        # Цикл обхода ошибок при определении первого столбца и отправителя
        while True:
            try:
                a=sheet.cell(n, 18).value
                track=sheet.cell(n, 4).value
            except Exception:
                print("Ошибка при определении трека и наличия ФИО")
                time.sleep(30)
                continue
            break

        # Если первый столбец с датой и столбец с треком не пустые то заполняем отправителя
        if a == None and track != None:
            print(rf"{n} заполняем")
            # # Цикл обхода ошибок при получении номера трека
            # while True:
            #     try:
            #         track = sheet.cell(n, 4).value  # Получаем трек
            #         print(type(track))
            #         if type(track)==str:
            #             print("Вместо трека слова - пропускаем")
            #             n+=1
            #             continue
            #     except Exception:
            #         print("Ошибка при получении трека из таблицы")
            #         time.sleep(15)
            #         continue
            #     break


            try:
                track=int(track)
            except Exception as e:
                print(rf"Ошибка при преобразовании трека в число {e.args}")
                n+=1
                continue

            browser.get(rf"https://www.pochta.ru/tracking#{str(track)}")  # Вводим трек в почту России
            browser.refresh()

            # Получаем ФИО отправителя

            # Словарь для возможных вариантов xpath
            xpath=['//*[@id="tracking-card-portal"]/div[4]/div[2]',
                   '//*[@id="tracking-card-portal"]/div[5]/div[2]',
                   ]

            # Цикл обхода ошибки при получении ФИО из почты России.
            for i in xpath:
                try:
                    FIO = wait.until(EC.visibility_of_element_located((By.XPATH, i))).text
                except Exception:
                    print("Ошибка почты России. Пробую другой XPATH")
                    # В случае ошибки перезапускаем браузер и снова открываем страницу с треком
                    browser.quit()
                    browser = webdriver.Chrome(driver_link, options=options)
                    browser.maximize_window()
                    wait = WebDriverWait(browser, 30)
                    browser.get(rf"https://www.pochta.ru/tracking#{str(track)}")  # Вводим трек в почту России
                    browser.refresh()
                    FIO="Ошибка"
                    continue
                break

            # Если по итогу обхода цикла ФИО не получилось вытащить, то идем в следующую строку
            if FIO=="Ошибка":
                print("Ошибка почты России, пропускаем")
                n+=1
                file = open('1.txt', 'w', encoding='utf-8')
                file.write(str(n))
                file.close()
                continue

            # Цикл обхода ошибки при форматировании
            for i in range(2):
                print(rf"{FIO} на итерации № {i}")
                try:
                    FIO = str(FIO).split(":")[1]  # Форматируем имя отправителя
                except Exception:
                    if i==0:
                        FIO = wait.until(EC.visibility_of_element_located(
                        (By.XPATH, '//*[@id="page-tracking"]/div/div[2]/div/div[2]/div[2]/div/div[4]/div[2]'))).text
                        continue
                    if i==1:
                        print("Ошибка при форматрировании")
                        FIO = "Ошибка"
                break


            # Если по итогу обхода цикла ФИО не получилось отформатировать, то идем в следующую строку
            if FIO=="Ошибка":
                n+=1
                file = open('1.txt', 'w', encoding='utf-8')
                file.write(str(n))
                file.close()
                continue

            # цикл обхода ошибки при вставке ФИО в таблицу
            while True:
                try:
                    sheet.update_cell(n, 18, FIO)
                except Exception:
                    print("Ошибка при вставке трека в таблицу")
                    time.sleep(15)
                    continue
                break
            browser.refresh()
        else:
            print(rf"{n} пропускаем")
            time.sleep(2)
        n+=1
        # file = open('1.txt', 'w', encoding='utf-8')
        # file.write(str(n))
        # file.close()


    browser.quit()  # Закрываем браузер по окончанию цикла
    print("Список закончился")

    # Прописываем начальную строку в файл для начала следующего цикла
    # file = open('1.txt', 'w', encoding='utf-8')
    # file.write(str(2))
    # file.close()

    """Заносим последнюю дату в файл Эксель"""
    sheet_xlsx[f'D{index}'] = datetime.now()
    # tab.close()
    tab.save(tab_name)
    print("Файл xlsx сохранен")


while True:
    try:
        FIO()
    except Exception as E:
        print(E.args)

