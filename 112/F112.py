import openpyxl
from openpyxl import Workbook, load_workbook


#  Открываем список
spisok = load_workbook(r'C:\Users\User\Desktop\Python\112\Файлы\Список.xlsx', data_only=True)
print('Список открыт')

#  Получаем имена вкладок списка из лайвинформа
sheets_spisok = spisok.sheetnames
print(sheets_spisok)

#  Открываем шаблон Ф112
f112 = load_workbook(r'C:\Users\User\Desktop\Python\112\Файлы\Ф112ЭП.xlsx')
print('Заявление открыто')
sheet_f112 = f112.active

#  Цикл обхода всех вкладок файла список
for sheet_spisok in sheets_spisok:
    #  Определяем рабочую вкладку
    worksheet_spisok = spisok[sheet_spisok]
    print(worksheet_spisok)

    #  Запускаем цикл обхода строк в рабочей вкладке списка из Лайва
    n=1
    while str(worksheet_spisok[f'A{n}'].value) != 'None':
        print(str(n)+". "+worksheet_spisok[f'A{n}'].value)

        # Определяем ФИО
        FIO=str(worksheet_spisok[f'A{n}'].value)

        # Вписываем ФИО
        sheet_f112['K20']= str(FIO)

        # Определяем цену
        cena=str(worksheet_spisok[f'B{n}'].value)

        # Вписываем цену
        sheet_f112['E5']= str(cena)

        # Определяем цену прописью
        cena_prop=str(worksheet_spisok[f'C{n}'].value)

        # Вписываем цену прописью
        sheet_f112['AJ4']= cena_prop

        #  Сохраняем заявление
        f112.save(rf'C:\Users\User\Desktop\Python\112\Файлы\Готов\{str(n)}.xlsx')

        n+=1
    print(f'Цикл обхода строк вкладки {worksheet_spisok} закончен')

print('Цикл обхода вкладок закончен')

spisok.close()
print('Список закрыт')

f112.close()
print('Бланк закрыт')


