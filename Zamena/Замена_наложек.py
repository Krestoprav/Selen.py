import win32com
from win32com import client
import openpyxl
from openpyxl import Workbook, load_workbook
import PyPDF2
from PyPDF2 import PdfFileWriter, PdfFileReader, PdfFileMerger
import os
import win32api
import win32print

n=0

# Создаем экземпляр класса входного (чтение) и выходного файла (запись)
input2 = PdfFileReader(open("1.pdf", "rb"))
output = PdfFileWriter()

# Определяем и выводим количество страниц входного файла из почты России
numPages = input2.getNumPages()
print(rf"В документе {numPages} страниц")

# Запускаем цикл обхода всех страниц и создания файлов
for n in range(2):
    print(n)
    #  Открываем список
    spisok = load_workbook(r'C:\Users\User\Desktop\Python\Замена\Список.xlsx', data_only=True)
    print('Список открыт')

    #  Получаем имена вкладок списка из лайвинформа
    sheets_spisok = spisok.sheetnames
    print(sheets_spisok)

    #  Открываем шаблон Ф112
    f112 = load_workbook(r'C:\Users\User\Desktop\Python\Замена\Ф112ЭП-бланк.xlsx')
    print('Бланк открыт')
    sheet_f112 = f112.active

    #  Цикл обхода всех вкладок файла список
    for sheet_spisok in sheets_spisok:
        #  Определяем рабочую вкладку
        worksheet_spisok = spisok[sheet_spisok]
        # print(worksheet_spisok)

        #  Запускаем цикл обхода строк в рабочей вкладке списка из Лайва

        while str(worksheet_spisok[f'A{n+1}'].value) != 'None':
            # print(str(n)+". "+worksheet_spisok[f'A{n}'].value)

            # Определяем ФИО
            FIO=str(worksheet_spisok[f'A{n+1}'].value)

            # Вписываем ФИО
            sheet_f112['K20']= str(FIO)

            # Определяем цену
            cena=str(worksheet_spisok[f'B{n+1}'].value)

            # Вписываем цену
            sheet_f112['E5']= str(cena)

            # Определяем цену прописью
            cena_prop=str(worksheet_spisok[f'C{n+1}'].value)

            # Вписываем цену прописью
            sheet_f112['AJ4']= cena_prop

            #  Сохраняем заполоненный бланк
            f112.save(r'C:\Users\User\Desktop\Python\Замена\Ф112ЭП.xlsx')

            break  # Останавливаем цикл. Пока оставим цикл, может пригодится

        # print(f'Цикл обхода строк вкладки {worksheet_spisok} закончен')

    # print('Цикл обхода вкладок закончен')

    spisok.close()
    print('Список закрыт')

    f112.close()
    print('Бланк закрыт')

    # Open Microsoft Excel
    excel = client.Dispatch("Excel.Application")
    print("Win32 открыл эксель")
    # excel.Visible = False

    # Read Excel File
    sheets = excel.Workbooks.Open(rf'C:\Users\User\Desktop\Python\Замена\Ф112ЭП.xlsx')
    work_sheets = sheets.Worksheets[0]
    print("Win32 прочитал файл с бланком")

    # Convert into PDF File
    work_sheets.ExportAsFixedFormat(0, rf'C:\Users\User\Desktop\Python\Замена\Ф112ЭП{n}.pdf')
    print("Win32 сохранил бланк в PDF")

    # Закрыть эксель
    sheets.Close(True)

    # Создаем экземпляр класса входного (чтение) и выходного файла (запись)
    input1 = PdfFileReader(open(rf"Ф112ЭП{n}.pdf", "rb"))

    page1=input1.getPage(0)


    # обрезаем n-ую страницу входного файла
    print(n)
    page2 = input2.getPage(n)
    print(page2.cropBox.getLowerLeft())
    print(page2.cropBox.getLowerRight())
    print(page2.cropBox.getUpperLeft())
    print(page2.cropBox.getUpperRight())

    # print(page.mediaBox.getUpperRight_x(), page.mediaBox.getUpperRight_y())
    page2.trimBox.lowerLeft = (422, 0)
    page2.trimBox.upperRight = (841.889, 595.275)
    page2.cropBox.lowerLeft = (422, 0)
    page2.cropBox.upperRight = (841.889, 595.275)
    # page.rotateClockwise(90)

    output.addPage(page2)  # Добавляем в выходной файл адресный ярлык
    output.addPage(page1)  # Добавляем в выходной файл наложку


with open("out.pdf", "wb") as out_f:  # Сохраняем выходной файл
    output.write(out_f)

# Печать две страницы на доном листе

name = win32print.GetDefaultPrinter()

#printdefaults = {"DesiredAccess": win32print.PRINTER_ACCESS_ADMINISTER}
printdefaults = {"DesiredAccess": win32print.PRINTER_ACCESS_USE}
handle = win32print.OpenPrinter(name, printdefaults)

level = 2
attributes = win32print.GetPrinter(handle, level)

print("Old Duplex = %d" % attributes['pDevMode'].Duplex)

#attributes['pDevMode'].Duplex = 1    # no flip
#attributes['pDevMode'].Duplex = 2    # flip up
attributes['pDevMode'].Duplex = 3    # flip over

## 'SetPrinter' fails because of 'Access is denied.'
## But the attribute 'Duplex' is set correctly
try:
    win32print.SetPrinter(handle, level, attributes, 0)
except:
    print("win32print.SetPrinter: set 'Duplex'")

res = win32api.ShellExecute(0, 'print', rf'C:\Users\User\Desktop\Python\Замена\out.pdf', None, '.', 0)

win32print.ClosePrinter(handle)

