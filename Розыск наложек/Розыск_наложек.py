import openpyxl
from openpyxl import Workbook, load_workbook, formatting
import win32com.client


# Словарь для ФИО отправителя
slovar_fio = {'Козырева': 'Козыреву Анастасию Сергеевну',
              'Рудак': 'Рудака Вячеслава Евгеньевича'

              }

#  Открываем список из Лавинформа
spisok = load_workbook(r'C:\Users\User\Desktop\Python\Розыск наложек\Файлы\Список.xlsx')
print('Список открыт')

#  Получаем имена вкладок списка из лайвинформа
sheets_spisok = spisok.sheetnames
print(sheets_spisok)

#  Открываем шаблон заявления
zaiavlenie = load_workbook(r'C:\Users\User\Desktop\Python\Розыск наложек\Файлы\Заявление.xlsx')
print('Заявление открыто')
sheet_zaiavlenie = zaiavlenie.active

#  Цикл обхода всех вкладок списка из Лайвинформа
for sheet_spisok in sheets_spisok:
    #  Определяем рабочую вкладку
    worksheet_spisok = spisok[sheet_spisok]
    print(worksheet_spisok)

    #  Запускаем цикл обхода строк в рабочей вкладке списка из Лайва
    n=1
    while str(worksheet_spisok[f'A{n}'].value) != 'None':
        print(worksheet_spisok[f'A{n}'].value)
        #  Если статус возврат в пункт или прием, то работаем. Иначе пропускаем строку
        if worksheet_spisok[f'R{n}'].value == 'Приём' or worksheet_spisok[f'R{n}'].value == 'Возврат в пункт приёма':
            print(f"{n} Рабочая строка")

            #  Определяем номер перевода
            nomer_perevoda=str(worksheet_spisok[f'Q{n}'].value)

            #  Определяем трек
            rpo=str(worksheet_spisok[f'D{n}'].value)

            #  Определяем индекс
            index_from=str(worksheet_spisok[f'T{n}'].value)
            index_to=str(worksheet_spisok[f'U{n}'].value)

            if index_to == '143900' or index_to == '101000' or index_to == '141020' or index_to == '143913':
                index = index_from
            else:
                index = index_to

            #  Определяем ФИО
            fio=slovar_fio[str(worksheet_spisok[f'E{n}'].value)]

            #  Вписываем информацию в заявление
            sheet_zaiavlenie['A11']=f'       Прошу Вас дослать перевод №{nomer_perevoda} (РПО {rpo}) ' \
                                    f'с ОПС №{index} на ОПС №101000 на ' \
                                    f'{fio}. Доверенность прилагается.'

            #  Сохраняем заявление
            zaiavlenie.save(rf'C:\Users\User\Desktop\Python\Розыск наложек\Файлы\{str(nomer_perevoda)}.xlsx')
            pythoncom.CoInitialize()
            Excel = win32com.client.Dispatch("Excel.Application")
            Excel.Visible = 0
            wb1 = Excel.Workbooks.Open(rf'C:\Users\User\Desktop\Python\Розыск наложек\Файлы\{str(nomer_perevoda)}.xlsx')
            wb1.ExportAsFixedFormat(0, rf'C:\Users\User\Desktop\Python\Розыск наложек\Файлы\{str(nomer_perevoda)}.xlsx')
            wb1.Close()
        else:
            print(f"{n} Неподходящая строка")

        n+=1
    print(f'Цикл обхода вкладки {worksheet_spisok} закончен')

print('Цикл обхода вкладок закончен')

spisok.close()
print('Список закрыт')

zaiavlenie.close()
print('Заявление закрыто')


